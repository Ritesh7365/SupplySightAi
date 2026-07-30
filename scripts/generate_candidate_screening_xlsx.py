"""
Generate Backend Engineer (AI Platform) candidate screening Excel workbook.
Data sourced from publicly available LinkedIn/web profile snapshots.
"""

from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Candidate Screening"

headers = [
    "S.No",
    "Candidate Name",
    "LinkedIn URL",
    "Current Company",
    "Experience",
    "Current Location",
    "Current Designation",
    "Relevant Skills",
    "Screening Remarks",
    "Fit Score (1-5)",
]

thin = Border(
    left=Side(style="thin", color="D0D5DD"),
    right=Side(style="thin", color="D0D5DD"),
    top=Side(style="thin", color="D0D5DD"),
    bottom=Side(style="thin", color="D0D5DD"),
)
header_fill = PatternFill("solid", fgColor="0F3D68")
header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(wrap_text=True, vertical="top", horizontal="center")
alt_fill = PatternFill("solid", fgColor="F2F4F7")
note_fill = PatternFill("solid", fgColor="FFF4CE")

for col, h in enumerate(headers, 1):
    cell = ws.cell(1, col, h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    cell.border = thin

candidates = [
    {
        "name": "Khushbu Agrawal",
        "linkedin": "https://www.linkedin.com/in/khushbu-agrawal-58275a169",
        "company": "Brevo (Customer Engagement / Martech SaaS)",
        "exp": "~5.2–5.5 years",
        "location": "Greater Delhi Area / Noida (Open to NCR)",
        "designation": "Senior Software Engineer / Senior Backend Engineer",
        "skills": (
            "Go, Node.js, Kafka, Redis, MongoDB, ClickHouse, PostgreSQL, "
            "Microservices, Event Streaming, Kubernetes, Observability "
            "(Datadog/Sentry/Grafana), GenAI tooling"
        ),
        "remarks": (
            "Strongest JD match. Built production Kafka event ingestion at Brevo "
            "(email/WhatsApp/campaign platform). Hands-on Redis caching, scheduler/"
            "workflow reliability, and cloud-native ops. ~5 YOE fits band; Delhi NCR "
            "based; communication-platform + AI tooling exposure."
        ),
        "score": 5,
    },
    {
        "name": "Aftab Khan",
        "linkedin": "https://www.linkedin.com/in/aftabskhan",
        "company": "Primathon (GoKwik ecosystem) | Ex-Samsung",
        "exp": "~2–3 years (slightly below band; high signal)",
        "location": "Greater Delhi Area",
        "designation": "Backend Engineer (SDE II)",
        "skills": (
            "Node.js, Python, Kafka, AWS, Elasticsearch, Postgres, Redis, "
            "Distributed Systems, LLM/recommendation (LangChain), Messaging sync "
            "(SMS/MMS/RCS)"
        ),
        "remarks": (
            "Delhi NCR + AI personalization + communication stack. Samsung messaging "
            "sync and RCS/SMS work maps to communication infrastructure. Built LLM "
            "recommendation flows. YOE is light vs 3–5 preference; still strong for "
            "AI-platform backend interviews."
        ),
        "score": 4,
    },
    {
        "name": "Chirag Khanna",
        "linkedin": "https://www.linkedin.com/in/chirag-khanna-b643a0101",
        "company": "Nielsen | Ex-Paytm (commerce/ONDC)",
        "exp": "~4–6 years (estimated from multi-company backend tenure)",
        "location": "Noida / Gurugram (Delhi NCR)",
        "designation": "Engineering / Backend (Node.js microservices)",
        "skills": (
            "Node.js, Express, Java Spring Boot, Kafka, Redis, MongoDB, MySQL, "
            "AWS Lambda/EC2, Kubernetes, WebSockets, SQS, Microservices"
        ),
        "remarks": (
            "NCR-based production backend across high-traffic systems (Paytm ONDC, "
            "live-class WebSockets, Kafka/Redis). Strong event-driven + cloud fit. "
            "Product company exposure (Paytm/Nielsen) aligns with SaaS/product "
            "preference over pure services."
        ),
        "score": 4,
    },
    {
        "name": "Ritik Agarwal",
        "linkedin": "https://www.linkedin.com/in/ritik-agarwal-164856184",
        "company": "TCGRE",
        "exp": "~3+ years",
        "location": "India (verify city on LinkedIn; Real-estate/Fintech product)",
        "designation": "Software Development Engineer II",
        "skills": (
            "NestJS, Node.js, Apache Kafka, Microservices, MongoDB, PostgreSQL, "
            "Redis, Event-driven campaign execution, Real-time analytics"
        ),
        "remarks": (
            "Clear production ownership of Kafka-based event-driven microservices "
            "for multi-channel campaign execution and analytics — closest to "
            "triggers/workflows/orchestration in the JD. YOE in target band. Confirm "
            "Delhi NCR location preference on screen call."
        ),
        "score": 4,
    },
    {
        "name": "Aditya G.",
        "linkedin": "https://www.linkedin.com/in/adityagread",
        "company": "miniOrange (SaaS / Identity & Access / Cloud Security product)",
        "exp": "~3–5 years (profile cites scaling to 10M+ daily requests)",
        "location": "New Delhi, Delhi",
        "designation": "Software Engineer",
        "skills": (
            "Golang, Kafka, Redis, ClickHouse, Distributed Systems, GraphQL APIs, "
            "Event ingestion pipelines, Auditing/streaming architecture"
        ),
        "remarks": (
            "Delhi-based Go engineer with Kafka streaming + Redis caching on a SaaS "
            "product handling multi-million daily requests. Strong distributed-systems "
            "and low-latency API evidence. Validate exact YOE and AWS depth in screen."
        ),
        "score": 4,
    },
    {
        "name": "Vaibhav Garg",
        "linkedin": "https://www.linkedin.com/in/vaibhavgarg3210",
        "company": "VVDN Technologies | Prior product/banking microservices work",
        "exp": "~4 years",
        "location": "Ghaziabad / Noida (Delhi NCR)",
        "designation": "Senior Software Engineer / Java Backend Developer",
        "skills": (
            "Java, Spring Boot, Microservices, REST APIs, Kafka, Docker, Eureka, "
            "API Gateway, Resilience4j retries/circuit-breaker, Observability"
        ),
        "remarks": (
            "Exact YOE band and NCR location. Built fault-tolerant event-driven "
            "microservices with retries, gateway auth, and Kafka — maps to "
            "triggers/scheduling/retries/workflows. Less AI/personalization signal; "
            "strong core backend + reliability hire."
        ),
        "score": 3,
    },
    {
        "name": "Piyush Kakkar",
        "linkedin": "https://www.linkedin.com/in/piyushk96",
        "company": "iion | Ex-Collegedunia (product; attribution/ad platforms)",
        "exp": "~7–8 years (above band; high seniority)",
        "location": "Delhi / Gurugram (Delhi NCR)",
        "designation": "Lead Backend Engineer",
        "skills": (
            "Go, Node.js/NestJS, Java, Kafka, RabbitMQ, SQS, MongoDB, Redis, AWS "
            "(EC2, Lambda, API Gateway, SQS), GCP Pub/Sub, GenAI, Kubernetes, "
            "Observability"
        ),
        "remarks": (
            "Exceptional stack match (queues, AWS, Mongo/Redis, GenAI) and NCR "
            "location. Scaled attribution platform toward billions of requests/day "
            "and real-time bidding. YOE above 3–5; consider only if open to Lead "
            "leveling or compensation stretch."
        ),
        "score": 4,
    },
    {
        "name": "Sai Praveen",
        "linkedin": "https://www.linkedin.com/in/pr0v33n",
        "company": "MoEngage Inc. (Customer Engagement Platform)",
        "exp": "~4.5+ years at MoEngage (total backend tenure aligns ~4–6)",
        "location": "Hyderabad (relocation / remote to be confirmed)",
        "designation": "Software Engineer – Product Analytics Backend",
        "skills": (
            "Python, Kafka, AWS, MongoDB, Airflow, Spark, FastAPI/Flask/Django, "
            "Analytics pipelines, Scalable backend product development"
        ),
        "remarks": (
            "Direct customer-engagement platform experience (MoEngage) — preferred "
            "company type. Builds analytics/backend for engagement product with "
            "Kafka/AWS/Mongo. Location is Hyderabad (not Delhi NCR); prioritize if "
            "remote/hybrid NCR or relocation is acceptable."
        ),
        "score": 4,
    },
    {
        "name": "Rahul Batheja",
        "linkedin": "https://www.linkedin.com/in/rahul-batheja",
        "company": "CleverTap (Customer Engagement / Communication SaaS)",
        "exp": "~8+ years (above band; Staff-level)",
        "location": "India (CleverTap hubs include Gurgaon — confirm city)",
        "designation": "Staff Backend Engineer / Team Lead Backend (CleverTap)",
        "skills": (
            "Java, Spring Boot, Kafka, Segmentation pipelines, Analytical DB "
            "optimization, Event-driven systems at billion-event scale, Mentoring"
        ),
        "remarks": (
            "Ideal domain: CleverTap segmentation/journeys at billions of events/day. "
            "Deep communication + personalization infrastructure. Seniority exceeds "
            "3–5 YOE JD; use as stretch/architect-track candidate or referral source "
            "for mid-level peers."
        ),
        "score": 4,
    },
    {
        "name": "Yash Vardhan Maurya",
        "linkedin": "https://www.linkedin.com/in/yash-vardhan-maurya-205311199",
        "company": "Info Edge India Ltd (Naukri / product SaaS group)",
        "exp": "~2.5–3.5 years (borderline lower band)",
        "location": "Noida, Uttar Pradesh (Delhi NCR)",
        "designation": "Software Developer (Backend-focused)",
        "skills": (
            "Java, Spring Boot, Kafka, Redis, MySQL, Elasticsearch, Node.js, "
            "MongoDB, Distributed Systems"
        ),
        "remarks": (
            "NCR + product company (Info Edge). Kafka/Redis/ES backend focus matches "
            "event/data path needs. YOE may sit at lower end of band; good pipeline "
            "candidate if screen shows production ownership depth."
        ),
        "score": 3,
    },
]

for i, c in enumerate(candidates, 1):
    row = i + 1
    values = [
        i,
        c["name"],
        c["linkedin"],
        c["company"],
        c["exp"],
        c["location"],
        c["designation"],
        c["skills"],
        c["remarks"],
        c["score"],
    ]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row, col, val)
        cell.alignment = center if col in (1, 5, 10) else wrap
        cell.border = thin
        cell.font = Font(name="Calibri", size=10)
        if i % 2 == 0:
            cell.fill = alt_fill
        if col == 3:
            cell.font = Font(name="Calibri", size=9, color="0563C1", underline="single")

widths = [6, 22, 48, 42, 28, 28, 36, 55, 55, 12]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 30
for r in range(2, 12):
    ws.row_dimensions[r].height = 95
ws.auto_filter.ref = "A1:J11"
ws.freeze_panes = "B2"

# Role brief
ws2 = wb.create_sheet("Role Brief")
ws2["A1"] = "Backend Engineer (AI Platform & Communication Infrastructure) — Sourcing Summary"
ws2["A1"].font = Font(bold=True, size=14, color="0F3D68")
brief = [
    ("Role", "Backend Engineer (AI Platform & Communication Infrastructure)"),
    ("Location", "Delhi NCR"),
    ("Experience Band", "3–5 Years"),
    ("Employment Type", "Full-time"),
    (
        "Must-have stack themes",
        "Node.js / Go / Python; REST; Distributed Systems; Event-driven Architecture; Message Queues; DBs; AWS",
    ),
    (
        "Preferred company types",
        "AI Product, D2C Product, SaaS, Customer Engagement, Communication Infrastructure, High-growth startups",
    ),
    (
        "Good-to-have",
        "AI/ML infra, personalization/recommendations, WhatsApp/Email/Push/SMS platforms, 0→1, experimentation",
    ),
    ("Sourcing date", str(date.today())),
    (
        "Sourcer note",
        "Profiles sourced from publicly available LinkedIn/web snapshots. Always re-verify live LinkedIn before outreach.",
    ),
]
ws2["A3"] = "Field"
ws2["B3"] = "Value"
ws2["A3"].font = Font(bold=True, color="FFFFFF")
ws2["B3"].font = Font(bold=True, color="FFFFFF")
ws2["A3"].fill = header_fill
ws2["B3"].fill = header_fill
for i, (k, v) in enumerate(brief, 4):
    ws2.cell(i, 1, k).font = Font(bold=True)
    ws2.cell(i, 2, v).alignment = wrap
    ws2.cell(i, 1).border = thin
    ws2.cell(i, 2).border = thin
ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["B"].width = 100

# Requester LinkedIn
ws3 = wb.create_sheet("Requester LinkedIn Access")
ws3["A1"] = "Your LinkedIn profile (as provided)"
ws3["A1"].font = Font(bold=True, size=13, color="0F3D68")
ws3["A3"] = "LinkedIn URL"
ws3["B3"] = "https://www.linkedin.com/in/ritesh-shinde-2483292b8/"
ws3["A4"] = "Automated access status"
ws3["B4"] = "BLOCKED — LinkedIn returned HTTP 999 / auth wall. Profile HTML could not be scraped."
ws3["B4"].fill = note_fill
ws3["A6"] = "Candidate Name"
ws3["B6"] = "Ritesh Shinde"
ws3["A7"] = "Current Company"
ws3["B7"] = "<< PASTE FROM YOUR LINKEDIN >>"
ws3["A8"] = "Experience"
ws3["B8"] = "<< PASTE YEARS OF EXPERIENCE >>"
ws3["A9"] = "Current Location"
ws3["B9"] = "<< PASTE LOCATION >>"
ws3["A10"] = "Current Designation"
ws3["B10"] = "<< PASTE HEADLINE / TITLE >>"
ws3["A11"] = "Relevant Skills"
ws3["B11"] = "<< PASTE SKILLS FROM PROFILE >>"
ws3["A12"] = "Screening Remarks"
ws3["B12"] = (
    "If this profile should be Candidate #1 for the role, paste fields above "
    "and reply in chat — the main screening sheet can be updated."
)
for r in range(3, 13):
    ws3.cell(r, 1).font = Font(bold=True)
    ws3.cell(r, 1).border = thin
    ws3.cell(r, 2).border = thin
    ws3.cell(r, 2).alignment = wrap
ws3.column_dimensions["A"].width = 28
ws3.column_dimensions["B"].width = 90
ws3.row_dimensions[12].height = 60
ws3["A14"] = "How to complete this sheet"
ws3["A14"].font = Font(bold=True, size=12)
ws3["A15"] = (
    "1) Open your LinkedIn while logged in.\n"
    "2) Copy Headline, Experience, Location, About/Skills.\n"
    "3) Paste into the cells above OR reply in chat with those details.\n"
    "4) Re-open each candidate LinkedIn URL before outreach; public snippets can lag."
)
ws3["A15"].alignment = wrap
ws3.merge_cells("A15:B18")

# Screening checklist
ws4 = wb.create_sheet("Screening Checklist")
ws4["A1"] = "Phone-screen checklist (use for each shortlisted candidate)"
ws4["A1"].font = Font(bold=True, size=13, color="0F3D68")
checks = [
    "Describe a production event-driven system you owned (producers, consumers, retries, DLQ).",
    "How did you design low-latency APIs and measure p95/p99?",
    "Experience with Kafka/Kinesis/SQS — partitioning, ordering, idempotency.",
    "AWS services used in production (Lambda, API Gateway, EC2, SQS) and failure modes handled.",
    "Any personalization, ranking, experimentation, or ML model serving experience?",
    "Communication channels worked on (Email/Push/SMS/WhatsApp) and delivery/orchestration patterns.",
    "Notice period, CTC expectations, willingness for Delhi NCR / hybrid.",
    "College background (Tier-1 preference) — verify if relevant to hiring policy.",
]
ws4["A3"] = "#"
ws4["B3"] = "Question / Check"
ws4["A3"].fill = header_fill
ws4["B3"].fill = header_fill
ws4["A3"].font = Font(bold=True, color="FFFFFF")
ws4["B3"].font = Font(bold=True, color="FFFFFF")
for i, q in enumerate(checks, 1):
    ws4.cell(i + 3, 1, i).border = thin
    ws4.cell(i + 3, 2, q).border = thin
    ws4.cell(i + 3, 2).alignment = wrap
ws4.column_dimensions["A"].width = 6
ws4.column_dimensions["B"].width = 100

out = r"D:\Data Science Project\SupplySightAi\Backend_Engineer_AI_Platform_Candidate_Screening.xlsx"
wb.save(out)
print("Saved:", out)
print("Candidates:", len(candidates))
