"""
Rebuild candidate screening Excel — Tier-1 institutes only (IIT / NIT / IIIT / BITS).
"""

from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Tier1 Candidate Screening"

headers = [
    "S.No",
    "Candidate Name",
    "LinkedIn URL",
    "Current Company",
    "Experience",
    "Current Location",
    "Current Designation",
    "Tier-1 Institute",
    "Relevant Skills",
    "Screening Remarks",
    "Fit Score (1-5)",
]

thin = Border(
    left=Side(style="thin", color="D0D5DD"),
    right=Side(style="thin", color="D0D5DD"),
    top=Side(style="thin", color="D0D5DD"),
    bottom=Side(style="thin", color="D0D5DD"),
)
header_fill = PatternFill("solid", fgColor="0F3D68")
header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(wrap_text=True, vertical="top", horizontal="center")
alt_fill = PatternFill("solid", fgColor="F2F4F7")
tier_fill = PatternFill("solid", fgColor="E8F5E9")

for col, h in enumerate(headers, 1):
    cell = ws.cell(1, col, h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    cell.border = thin

candidates = [
    {
        "name": "Abhinav Kaushik",
        "linkedin": "https://www.linkedin.com/in/abhinav2578",
        "company": "Atlys (product / travel-tech SaaS)",
        "exp": "~3.9 years",
        "location": "New Delhi, Delhi (Delhi NCR)",
        "designation": "Senior Software Engineer – Backend",
        "institute": "IIT Delhi (B.Tech)",
        "skills": (
            "Python, Node.js, Django, Sanic, AWS, Docker, Kubernetes, "
            "PostgreSQL, Redis, Kafka, SQS, Microservices, REST APIs"
        ),
        "remarks": (
            "Strong Tier-1 + NCR + stack match. IIT Delhi with production backend "
            "on Node/Python, Kafka, Redis, SQS, AWS. Fits 3–5 YOE band and "
            "distributed/event-driven JD requirements. Product-company background."
        ),
        "score": 5,
    },
    {
        "name": "Aditya",
        "linkedin": "https://www.linkedin.com/in/aditya-36583a163",
        "company": "ArmorCode Inc. | Ex-Samsung R&D",
        "exp": "~5.5 years (upper end of band / slight stretch)",
        "location": "Gurugram, Haryana (Delhi NCR) | Ex-Noida at Samsung",
        "designation": "Senior Software Engineer / Backend Engineer",
        "institute": "IIT Bombay (M.Tech Computer Science)",
        "skills": (
            "Java, Spring Boot, Kafka, Redis, Distributed Systems, AWS, "
            "Microservices, Ex-Samsung production systems"
        ),
        "remarks": (
            "IIT Bombay + Delhi NCR. Kafka/Redis distributed-systems depth and "
            "Samsung R&D Noida experience. Slightly above 5 YOE; still strong for "
            "AI-platform/backend reliability hiring."
        ),
        "score": 5,
    },
    {
        "name": "Vaibhav Kashera",
        "linkedin": "https://www.linkedin.com/in/vaibhavkashera",
        "company": "Sprinklr (Customer Engagement / Social & CX platform)",
        "exp": "~2–3 years post IIIT-H (SDE-II level; verify exact YOE)",
        "location": "Gurugram, Haryana (Delhi NCR)",
        "designation": "SDE-II (Backend)",
        "institute": "IIIT Hyderabad (B.Tech + MS CSE)",
        "skills": (
            "Backend engineering, Distributed systems, Platform engineering, "
            "Customer engagement product stack (Sprinklr)"
        ),
        "remarks": (
            "IIIT Hyderabad (Tier-1) + Sprinklr — direct customer-engagement "
            "platform preference in JD. Gurugram-based. Confirm Kafka/AWS depth "
            "and exact YOE on screen; domain fit is excellent."
        ),
        "score": 5,
    },
    {
        "name": "Sai Praveen",
        "linkedin": "https://www.linkedin.com/in/pr0v33n",
        "company": "MoEngage Inc. (Customer Engagement Platform)",
        "exp": "~4.5+ years",
        "location": "Hyderabad (relocation / remote for NCR TBD)",
        "designation": "Software Engineer – Product Analytics Backend",
        "institute": "IIT (ISM) Dhanbad (B.Tech CSE)",
        "skills": (
            "Python, Kafka, AWS, MongoDB, Airflow, Spark, FastAPI/Flask/Django, "
            "Analytics pipelines, Scalable product backend"
        ),
        "remarks": (
            "IIT + MoEngage is a top domain match for communication/engagement "
            "infrastructure. Kafka/AWS/Mongo production ownership. Location is "
            "Hyderabad — prioritize if open to relocate/hybrid NCR."
        ),
        "score": 5,
    },
    {
        "name": "Aayush Gupta",
        "linkedin": "https://www.linkedin.com/in/aayush-27",
        "company": "Telna | Ex-Mobileum | Ex-Deutsche Telekom Digital Labs",
        "exp": "~4.5 years",
        "location": "India (IIIT Delhi alum; verify current city for NCR)",
        "designation": "Software Development Engineer / Backend",
        "institute": "IIIT Delhi (B.Tech ECE)",
        "skills": (
            "Java, Go, Kafka, Redis, Elasticsearch, RabbitMQ, Microservices, "
            "REST APIs, A/B experimentation, Telecom/engagement systems"
        ),
        "remarks": (
            "IIIT Delhi + ~4.5 YOE in Java/Go backend with Kafka and experimentation. "
            "Telecom/digital labs background maps to communication infrastructure. "
            "Confirm Delhi NCR willingness and current location."
        ),
        "score": 4,
    },
    {
        "name": "Harshal Dev",
        "linkedin": "https://www.linkedin.com/in/harshaldev",
        "company": "Jio",
        "exp": "~3.4 years",
        "location": "Bengaluru (IIIT Delhi’23; notice-period signal on profile)",
        "designation": "Software Engineer 2 (SDE 2)",
        "institute": "IIIT Delhi (B.Tech CSE)",
        "skills": (
            "Java, Spring Boot, Kafka, Real-time analytics, Microservices, "
            "Distributed systems"
        ),
        "remarks": (
            "IIIT Delhi in exact YOE band with Kafka + real-time analytics. "
            "Serving notice (per profile) can mean faster joining. Location is "
            "Bangalore — check NCR relocate interest."
        ),
        "score": 4,
    },
    {
        "name": "Yash Chhaparia",
        "linkedin": "https://www.linkedin.com/in/yashchhaparia",
        "company": "CRED (fintech product)",
        "exp": "~3+ years",
        "location": "Kanpur / India (verify; open to SDE-2 backend roles)",
        "designation": "Software Engineer",
        "institute": "IIT Kanpur (B.Tech)",
        "skills": (
            "Java, Spring Boot, Kafka, MySQL/MongoDB, Redis, Temporal, AWS, "
            "High-throughput distributed transaction services"
        ),
        "remarks": (
            "IIT Kanpur + CRED product backend with Kafka/Temporal/AWS at scale "
            "(~30k peak RPM cited). Strong distributed-systems signal in 3–5 YOE "
            "band. Confirm Delhi NCR preference."
        ),
        "score": 4,
    },
    {
        "name": "Ayan Kumar Pahari",
        "linkedin": "https://www.linkedin.com/in/ayankumarpahari",
        "company": "Razorpay (fintech product)",
        "exp": "~3–4.8 years",
        "location": "Bengaluru (IIT Hyderabad M.Tech)",
        "designation": "Senior Software Development Engineer",
        "institute": "IIT Hyderabad (M.Tech CSE) | GATE/research background",
        "skills": (
            "Go, gRPC, Apache Kafka, Temporal, AWS, Docker, Kubernetes, "
            "Event-driven workflows, Distributed systems, AI hackathon exposure"
        ),
        "remarks": (
            "IIT Hyderabad + Razorpay: Go/Kafka/Temporal event-driven systems at "
            "enterprise scale. Excellent orchestration/workflow match for JD. "
            "Bangalore-based — assess NCR relocate/remote."
        ),
        "score": 4,
    },
    {
        "name": "Kushal Jindal",
        "linkedin": "https://www.linkedin.com/in/kushal-jindal-843102201",
        "company": "Zepto (D2C / quick-commerce product) | Recommendation work",
        "exp": "~2–3 years post IITR (verify; profile cites ~5y total incl. related)",
        "location": "Bengaluru / Patiala listed historically",
        "designation": "Backend Software Engineer",
        "institute": "IIT Roorkee (B.Tech Electrical)",
        "skills": (
            "Go, Kafka, AWS, Kubernetes, Docker, REST APIs, Event-driven systems, "
            "Recommendation / personalization (Zepto intern + backend)"
        ),
        "remarks": (
            "IIT Roorkee + Zepto D2C product preference. Event-driven Go/Kafka/AWS "
            "plus recommendation/personalization — rare JD good-to-have combo. "
            "Validate exact YOE vs 3–5 band and NCR interest."
        ),
        "score": 4,
    },
    {
        "name": "Yash Vardhan Maurya",
        "linkedin": "https://www.linkedin.com/in/yash-vardhan-maurya-205311199",
        "company": "Info Edge India Ltd (Naukri group / product SaaS)",
        "exp": "~2.5–3.5 years (lower end of band)",
        "location": "Noida, Uttar Pradesh (Delhi NCR)",
        "designation": "Software Developer (Backend-focused)",
        "institute": "IIIT Kota (B.Tech CSE) — Institute of National Importance",
        "skills": (
            "Java, Spring Boot, Kafka, Redis, MySQL, Elasticsearch, Node.js, "
            "MongoDB, Distributed Systems"
        ),
        "remarks": (
            "IIIT Kota (Tier-1 per JD IIIT list) + Noida product company. "
            "Kafka/Redis/ES backend focus. YOE at lower band — good pipeline if "
            "screen shows production ownership."
        ),
        "score": 4,
    },
]

for i, c in enumerate(candidates, 1):
    row = i + 1
    values = [
        i,
        c["name"],
        c["linkedin"],
        c["company"],
        c["exp"],
        c["location"],
        c["designation"],
        c["institute"],
        c["skills"],
        c["remarks"],
        c["score"],
    ]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row, col, val)
        cell.alignment = center if col in (1, 5, 11) else wrap
        cell.border = thin
        cell.font = Font(name="Calibri", size=10)
        if i % 2 == 0:
            cell.fill = alt_fill
        if col == 3:
            cell.font = Font(name="Calibri", size=9, color="0563C1", underline="single")
        if col == 8:
            cell.fill = tier_fill
            cell.font = Font(name="Calibri", size=10, bold=True)

widths = [6, 22, 46, 40, 28, 32, 34, 36, 50, 52, 12]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 32
for r in range(2, 12):
    ws.row_dimensions[r].height = 100
ws.auto_filter.ref = "A1:K11"
ws.freeze_panes = "B2"

# Summary sheet
ws2 = wb.create_sheet("Tier1 Summary")
ws2["A1"] = "Tier-1 Rebuild — Backend Engineer (AI Platform & Communication Infrastructure)"
ws2["A1"].font = Font(bold=True, size=14, color="0F3D68")
summary_lines = [
    ("Filter applied", "IIT / NIT / IIIT / BITS only (per JD mandatory preference)"),
    ("Previous shortlist Tier-1 count", "2 / 10"),
    ("New shortlist Tier-1 count", "10 / 10"),
    ("NCR-based / NCR-friendly highlights", "Abhinav Kaushik, Aditya, Vaibhav Kashera, Yash Vardhan Maurya"),
    ("Engagement / communication domain", "Vaibhav Kashera (Sprinklr), Sai Praveen (MoEngage)"),
    ("AI / personalization signal", "Kushal Jindal (recommendations), Vaibhav/Sai domain platforms"),
    ("Sourcing date", str(date.today())),
    (
        "Caveat",
        "Public LinkedIn snapshots; re-verify institute, YOE, location, and notice period before outreach.",
    ),
]
ws2["A3"] = "Field"
ws2["B3"] = "Value"
ws2["A3"].fill = header_fill
ws2["B3"].fill = header_fill
ws2["A3"].font = Font(bold=True, color="FFFFFF")
ws2["B3"].font = Font(bold=True, color="FFFFFF")
for i, (k, v) in enumerate(summary_lines, 4):
    ws2.cell(i, 1, k).font = Font(bold=True)
    ws2.cell(i, 1).border = thin
    ws2.cell(i, 2, v).alignment = wrap
    ws2.cell(i, 2).border = thin
ws2.column_dimensions["A"].width = 36
ws2.column_dimensions["B"].width = 95

# Institute checklist
ws3 = wb.create_sheet("Institute Verification")
ws3["A1"] = "Institute verification (Tier-1 gate)"
ws3["A1"].font = Font(bold=True, size=13, color="0F3D68")
ws3["A3"] = "Candidate"
ws3["B3"] = "Institute"
ws3["C3"] = "Category"
ws3["D3"] = "Passes Tier-1 filter?"
for col in range(1, 5):
    ws3.cell(3, col).fill = header_fill
    ws3.cell(3, col).font = Font(bold=True, color="FFFFFF")
    ws3.cell(3, col).border = thin

institute_rows = [
    ("Abhinav Kaushik", "IIT Delhi", "IIT", "Yes"),
    ("Aditya", "IIT Bombay (M.Tech)", "IIT", "Yes"),
    ("Vaibhav Kashera", "IIIT Hyderabad", "IIIT", "Yes"),
    ("Sai Praveen", "IIT (ISM) Dhanbad", "IIT", "Yes"),
    ("Aayush Gupta", "IIIT Delhi", "IIIT", "Yes"),
    ("Harshal Dev", "IIIT Delhi", "IIIT", "Yes"),
    ("Yash Chhaparia", "IIT Kanpur", "IIT", "Yes"),
    ("Ayan Kumar Pahari", "IIT Hyderabad (M.Tech)", "IIT", "Yes"),
    ("Kushal Jindal", "IIT Roorkee", "IIT", "Yes"),
    ("Yash Vardhan Maurya", "IIIT Kota", "IIIT (INI)", "Yes"),
]
for i, row in enumerate(institute_rows, 4):
    for col, val in enumerate(row, 1):
        cell = ws3.cell(i, col, val)
        cell.border = thin
        if col == 4:
            cell.fill = tier_fill
            cell.font = Font(bold=True, color="1B5E20")
ws3.column_dimensions["A"].width = 24
ws3.column_dimensions["B"].width = 28
ws3.column_dimensions["C"].width = 14
ws3.column_dimensions["D"].width = 22

# Screening checklist
ws4 = wb.create_sheet("Screening Checklist")
ws4["A1"] = "Phone-screen checklist (Tier-1 shortlist)"
ws4["A1"].font = Font(bold=True, size=13, color="0F3D68")
checks = [
    "Confirm institute degree year and campus (IIT/NIT/IIIT/BITS).",
    "Describe a production event-driven system (Kafka/SQS): retries, DLQ, idempotency.",
    "Low-latency API ownership — p95/p99 measurement and improvements.",
    "AWS production services used (Lambda, API Gateway, EC2, SQS) and failure handling.",
    "Any personalization, ranking, experimentation, or ML serving experience?",
    "Communication channels (Email/Push/SMS/WhatsApp) or engagement-platform work?",
    "Current location + willingness for Delhi NCR / hybrid; notice period; CTC.",
    "Open to 0→1 / startup pace vs large product org?",
]
ws4["A3"] = "#"
ws4["B3"] = "Question / Check"
ws4["A3"].fill = header_fill
ws4["B3"].fill = header_fill
ws4["A3"].font = Font(bold=True, color="FFFFFF")
ws4["B3"].font = Font(bold=True, color="FFFFFF")
for i, q in enumerate(checks, 1):
    ws4.cell(i + 3, 1, i).border = thin
    ws4.cell(i + 3, 2, q).border = thin
    ws4.cell(i + 3, 2).alignment = wrap
ws4.column_dimensions["A"].width = 6
ws4.column_dimensions["B"].width = 100

out_primary = r"D:\Data Science Project\SupplySightAi\Backend_Engineer_Tier1_Candidate_Screening.xlsx"
out_legacy = r"D:\Data Science Project\SupplySightAi\Backend_Engineer_AI_Platform_Candidate_Screening.xlsx"
wb.save(out_primary)
print("Saved:", out_primary)
try:
    wb.save(out_legacy)
    print("Also updated:", out_legacy)
except Exception as exc:
    print("Could not overwrite open legacy file (close Excel if needed):", exc)
print("Candidates:", len(candidates))
